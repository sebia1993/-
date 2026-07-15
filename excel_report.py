from __future__ import annotations

import math
from datetime import datetime, timedelta, timezone
from typing import Any

from openpyxl.chart import LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter


KST = timezone(timedelta(hours=9), name="KST")
KST_NOTE = "시간 기준: 한국 표준시(KST)"
EXCEL_DATETIME_FORMAT = "yyyy-mm-dd hh:mm:ss"


def to_kst_excel_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        parsed = value
    else:
        text = "" if value is None else str(value).strip()
        if not text:
            return None
        try:
            parsed = datetime.fromisoformat(text)
        except ValueError:
            try:
                parsed = datetime.strptime(text, "%Y-%m-%d %H:%M:%S %z")
            except ValueError:
                return None

    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=KST)
    return parsed.astimezone(KST).replace(tzinfo=None)


def kst_filename_timestamp(value: Any) -> str:
    parsed = to_kst_excel_datetime(value)
    return parsed.strftime("%Y%m%d_%H%M%S") if parsed is not None else "unknown-time"


def nice_axis_maximum(value: float) -> float:
    safe_value = max(10.0, float(value) * 1.04)
    magnitude = 10 ** math.floor(math.log10(safe_value))
    normalized = safe_value / magnitude
    step = next(candidate for candidate in (1, 1.5, 2, 2.5, 5, 7.5, 10) if normalized <= candidate)
    return float(step * magnitude)


def chart_label_positions(indexes: list[int], values: list[float]) -> list[int]:
    if not indexes or len(indexes) != len(values):
        return []
    if len(indexes) <= 12:
        return list(range(len(indexes)))

    positions = {position for position, index in enumerate(indexes) if index % 5 == 0}
    positions.update({values.index(min(values)), values.index(max(values)), len(indexes) - 1})
    return sorted(positions)


def add_speed_chart(
    sheet,
    *,
    title: str,
    header_row: int,
    start_row: int,
    end_row: int,
    time_column: int,
    speed_column: int,
    average_column: int,
    label_column: int,
    anchor: str,
    color: str,
    indexes: list[int],
    values: list[float],
    average_value: float | None = None,
) -> LineChart:
    average = (
        float(average_value)
        if average_value is not None and math.isfinite(float(average_value)) and average_value >= 0
        else sum(values) / len(values)
    )
    minimum = min(values)
    maximum = max(values)
    sheet.cell(header_row, label_column, "표시값")
    for position in chart_label_positions(indexes, values):
        sheet.cell(start_row + position, label_column, values[position])

    chart = LineChart()
    chart.title = (
        f"{title}\n"
        f"평균 {average:.1f} · 최저 {minimum:.1f} · 최고 {maximum:.1f} Mbps"
    )
    chart.y_axis.title = "속도(Mbps)"
    chart.x_axis.title = "시간(초)"
    chart.height = 8.5
    chart.width = 17
    chart.legend.position = "b"
    chart.legend = None
    chart.visible_cells_only = False
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = nice_axis_maximum(max(maximum, average))
    chart.y_axis.majorUnit = chart.y_axis.scaling.max / 4
    chart.y_axis.numFmt = "0.0"

    categories = Reference(sheet, min_col=time_column, min_row=start_row, max_row=end_row)
    actual_data = Reference(sheet, min_col=speed_column, min_row=header_row, max_row=end_row)
    average_data = Reference(sheet, min_col=average_column, min_row=header_row, max_row=end_row)
    label_data = Reference(sheet, min_col=label_column, min_row=header_row, max_row=end_row)
    chart.add_data(actual_data, titles_from_data=True)
    chart.add_data(average_data, titles_from_data=True)
    chart.add_data(label_data, titles_from_data=True)
    chart.set_categories(categories)

    actual_series = chart.series[0]
    actual_series.val.numRef.f = (
        f"'{sheet.title}'!${get_column_letter(speed_column)}${start_row}:"
        f"${get_column_letter(speed_column)}${end_row}"
    )
    actual_series.graphicalProperties.line.solidFill = color
    actual_series.graphicalProperties.line.width = 26000
    actual_series.marker.symbol = "circle"
    actual_series.marker.size = 5
    actual_series.marker.graphicalProperties.solidFill = color
    actual_series.marker.graphicalProperties.line.solidFill = color
    average_series = chart.series[1]
    average_series.val.numRef.f = (
        f"'{sheet.title}'!${get_column_letter(average_column)}${start_row}:"
        f"${get_column_letter(average_column)}${end_row}"
    )
    average_series.graphicalProperties.line.solidFill = "70766F"
    average_series.graphicalProperties.line.width = 16000
    average_series.graphicalProperties.line.prstDash = "dash"
    average_series.marker.symbol = "none"

    label_series = chart.series[2]
    label_series.val.numRef.f = (
        f"'{sheet.title}'!${get_column_letter(label_column)}${start_row}:"
        f"${get_column_letter(label_column)}${end_row}"
    )
    label_series.graphicalProperties.line.noFill = True
    label_series.marker.symbol = "none"
    label_series.dLbls = DataLabelList(showVal=True, dLblPos="t", numFmt='0.0" Mbps"')

    sheet.add_chart(chart, anchor)
    return chart


__all__ = [
    "EXCEL_DATETIME_FORMAT",
    "KST_NOTE",
    "add_speed_chart",
    "chart_label_positions",
    "kst_filename_timestamp",
    "nice_axis_maximum",
    "to_kst_excel_datetime",
]
