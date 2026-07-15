import csv
import json
import time
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app import create_app
from network_measurement import NetworkMeasurementGate
from network_sustained import (
    SUSTAINED_LOG_FIELDS,
    SustainedCheckError,
    SustainedCheckManager,
    SustainedCheckSettings,
    summarize_intervals,
)
from sustained_excel import (
    EXCEL_MIME_TYPE,
    build_sustained_excel,
    build_sustained_excel_filename,
    safe_excel_text,
)
from tests.test_app import write_config


class FakeClock:
    def __init__(self) -> None:
        self.value = 0.0

    def __call__(self) -> float:
        return self.value

    def advance(self, seconds: float) -> None:
        self.value += seconds


def create_manager(tmp_path: Path, clock: FakeClock, *, ttl: float = 300) -> SustainedCheckManager:
    return SustainedCheckManager(
        log_path=tmp_path / "data" / "network_check_session_log.csv",
        results_root=tmp_path / "data" / "network_check_results",
        settings=SustainedCheckSettings(session_ttl_seconds=ttl, download_chunk_bytes=1024),
        clock=clock,
    )


def sample_excel_result(*, status: str = "success", error: str = "") -> dict:
    upload_intervals = [
        {"index": index, "duration_seconds": 1.0, "bytes_transferred": index * 1_000_000, "mbps": index * 8.0}
        for index in range(1, 31)
    ]
    download_intervals = [
        {"index": index, "duration_seconds": 1.0, "bytes_transferred": index * 2_000_000, "mbps": index * 16.0}
        for index in range(1, 31)
    ]
    return {
        "schema_version": 1,
        "session_id": "0123456789abcdef0123456789abcdef",
        "client_ip": "10.0.0.10",
        "started_at": "2026-07-14 14:30:00 +0900",
        "completed_at": "2026-07-14 14:31:06 +0900",
        "requested": {
            "direction": "full",
            "duration_seconds": 30,
            "warmup_seconds": 3.0,
            "stream_count": 4,
        },
        "http_latency": {"samples_ms": [1.1, 1.3, 1.5], "min_ms": 1.1, "median_ms": 1.3, "max_ms": 1.5},
        "directions": {
            "upload": {
                "bytes_transferred": 465_000_000,
                "actual_duration_seconds": 30.0,
                "average_mbps": 124.0,
                "median_mbps": 124.0,
                "min_mbps": 8.0,
                "max_mbps": 240.0,
                "variability_percent": 55.8,
                "intervals": upload_intervals,
            },
            "download": {
                "bytes_transferred": 930_000_000,
                "actual_duration_seconds": 30.0,
                "average_mbps": 248.0,
                "median_mbps": 248.0,
                "min_mbps": 16.0,
                "max_mbps": 480.0,
                "variability_percent": 55.8,
                "intervals": download_intervals,
            },
        },
        "status": status,
        "error": error,
        "result_url": "/network-check/sustained/results/0123456789abcdef0123456789abcdef.json",
        "excel_url": "/network-check/sustained/results/0123456789abcdef0123456789abcdef.xlsx",
    }


def test_summarize_intervals_calculates_stable_statistics():
    summary = summarize_intervals(
        byte_count=3_000_000,
        duration_seconds=3,
        interval_bytes=[1_000_000, 1_000_000, 1_000_000],
    )

    assert summary["average_mbps"] == 8.0
    assert summary["median_mbps"] == 8.0
    assert summary["min_mbps"] == 8.0
    assert summary["max_mbps"] == 8.0
    assert summary["variability_percent"] == 0.0


def test_summarize_intervals_uses_actual_client_sample_durations():
    summary = summarize_intervals(
        byte_count=3_000_000,
        duration_seconds=3,
        interval_bytes=[900_000, 1_100_000, 1_000_000],
        interval_durations=[0.9, 1.1, 1.0],
    )

    assert summary["average_mbps"] == 8.0
    assert summary["min_mbps"] == 8.0
    assert summary["max_mbps"] == 8.0
    assert [row["duration_seconds"] for row in summary["intervals"]] == [0.9, 1.1, 1.0]


def test_excel_report_contains_summary_intervals_and_chart():
    workbook = load_workbook(BytesIO(build_sustained_excel(sample_excel_result())))

    assert workbook.sheetnames == ["결과 요약", "속도 변화"]
    summary = workbook["결과 요약"]
    intervals = workbook["속도 변화"]
    assert summary["A1"].value == "HTTP 시간 측정 결과"
    assert summary["A2"].value == "시간 기준: 한국 표준시(KST)"
    assert summary["B4"].value == "성공"
    assert summary["D4"].value == "전체"
    assert summary["F4"].value == "10.0.0.10"
    assert summary["B5"].value.strftime("%Y-%m-%d %H:%M:%S") == "2026-07-14 14:30:00"
    assert summary["B6"].value == 30
    assert summary["F6"].value == 4
    assert summary["A11"].value == "업로드"
    assert summary["B11"].value == 124.0
    assert summary["C11"].value == 15.5
    assert summary["A12"].value == "다운로드"
    assert intervals["A4"].value == "업로드"
    assert intervals["B4"].value == 1
    assert intervals["C4"].value == 8.0
    assert intervals["C63"].value == 480.0
    assert intervals.freeze_panes == "A4"
    assert intervals.auto_filter.ref == "A3:D63"
    assert len(intervals._charts) == 2
    assert all(len(chart.series) == 3 for chart in intervals._charts)
    assert all(chart.y_axis.scaling.min == 0 for chart in intervals._charts)
    assert intervals._charts[0].series[2].dLbls.showVal is True
    assert all(intervals.column_dimensions[column].hidden for column in "EFGH")
    assert all(
        cell.value != "0123456789abcdef0123456789abcdef"
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
    )


def test_excel_report_neutralizes_formula_text_and_supports_failure():
    result = sample_excel_result(status="failure", error="=HYPERLINK(\"https://invalid\",\"open\")")
    result["client_ip"] = "  +cmd|' /C calc'!A0"

    workbook = load_workbook(BytesIO(build_sustained_excel(result)), data_only=False)
    summary = workbook["결과 요약"]

    assert summary["B4"].value == "실패"
    assert summary["B7"].value.startswith("'=")
    assert summary["F4"].value.startswith("'")
    assert summary["B7"].data_type == "s"
    assert safe_excel_text("@SUM(A1:A2)") == "'@SUM(A1:A2)"


def test_excel_filename_uses_kst_completion_time_without_session_id():
    filename = build_sustained_excel_filename(sample_excel_result())

    assert filename == "HTTP_시간측정_20260714_143106.xlsx"
    assert "01234567" not in filename


def test_excel_report_converts_offset_time_to_kst():
    result = sample_excel_result()
    result["started_at"] = "2026-07-14 05:30:00 +0000"
    result["completed_at"] = "2026-07-14 06:31:06 +0000"

    workbook = load_workbook(BytesIO(build_sustained_excel(result)))

    assert workbook["결과 요약"]["B5"].value.strftime("%Y-%m-%d %H:%M:%S") == "2026-07-14 14:30:00"
    assert build_sustained_excel_filename(result) == "HTTP_시간측정_20260714_153106.xlsx"


def test_manager_excludes_warmup_and_persists_result(tmp_path):
    clock = FakeClock()
    manager = create_manager(tmp_path, clock)
    session = manager.start_session(
        client_ip="10.0.0.10",
        direction="upload",
        duration_seconds=10,
        stream_count=1,
    )

    manager.begin_phase(session.session_id, session.client_ip, "upload", "warmup")
    manager.record_bytes(session.session_id, session.client_ip, "upload", 500_000)
    clock.advance(3.1)
    manager.begin_phase(session.session_id, session.client_ip, "upload", "measure")
    manager.record_bytes(session.session_id, session.client_ip, "upload", 1_000_000)
    clock.advance(1.1)
    manager.record_bytes(session.session_id, session.client_ip, "upload", 2_000_000)
    clock.advance(9)

    result = manager.complete(
        session.session_id,
        session.client_ip,
        {"latency_samples_ms": [2.0, 3.0, 4.0]},
    )

    assert result["directions"]["upload"]["bytes_transferred"] == 3_000_000
    assert result["http_latency"]["median_ms"] == 3.0
    result_path = tmp_path / "data" / "network_check_results" / f"{session.session_id}.json"
    assert json.loads(result_path.read_text(encoding="utf-8"))["status"] == "success"
    with (tmp_path / "data" / "network_check_session_log.csv").open(
        "r", encoding="utf-8-sig", newline=""
    ) as handle:
        rows = list(csv.DictReader(handle))
    assert rows[0]["bytes_transferred"] == "3000000"
    assert rows[0]["direction"] == "upload"


def test_manager_csv_partial_write_failure_rolls_back_log_and_json(tmp_path, monkeypatch):
    gate = NetworkMeasurementGate()
    manager = SustainedCheckManager(
        log_path=tmp_path / "data" / "network_check_session_log.csv",
        results_root=tmp_path / "data" / "network_check_results",
        measurement_gate=gate,
    )
    session = manager.start_session(
        client_ip="10.0.0.10",
        direction="full",
        duration_seconds=10,
        stream_count=1,
    )
    original_log = manager.log_path.read_bytes()
    original_writerow = csv.DictWriter.writerow
    call_count = 0

    def fail_second_writerow(writer, row):
        nonlocal call_count
        call_count += 1
        if call_count == 2:
            raise OSError("disk full")
        return original_writerow(writer, row)

    monkeypatch.setattr(csv.DictWriter, "writerow", fail_second_writerow)

    with pytest.raises(OSError, match="disk full"):
        manager.cancel(session.session_id, session.client_ip)

    assert manager.log_path.read_bytes() == original_log
    assert not (manager.results_root / f"{session.session_id}.json").exists()
    assert manager.active_session is None
    assert gate.is_available() is True


def test_manager_rejects_parallel_session_and_wrong_ip(tmp_path):
    clock = FakeClock()
    manager = create_manager(tmp_path, clock)
    session = manager.start_session(
        client_ip="10.0.0.10",
        direction="download",
        duration_seconds=10,
        stream_count=1,
    )

    with pytest.raises(SustainedCheckError) as conflict:
        manager.start_session(
            client_ip="10.0.0.11",
            direction="upload",
            duration_seconds=10,
            stream_count=1,
        )
    assert conflict.value.status_code == 409

    with pytest.raises(SustainedCheckError) as forbidden:
        manager.status(session.session_id, "10.0.0.11")
    assert forbidden.value.status_code == 403


def test_manager_expiry_releases_active_slot_and_records_failure(tmp_path):
    clock = FakeClock()
    manager = create_manager(tmp_path, clock, ttl=5)
    session = manager.start_session(
        client_ip="10.0.0.10",
        direction="upload",
        duration_seconds=10,
        stream_count=1,
    )
    clock.advance(6)

    replacement = manager.start_session(
        client_ip="10.0.0.11",
        direction="download",
        duration_seconds=10,
        stream_count=1,
    )

    assert replacement.session_id != session.session_id
    expired_path = tmp_path / "data" / "network_check_results" / f"{session.session_id}.json"
    assert json.loads(expired_path.read_text(encoding="utf-8"))["status"] == "failure"


def test_manager_automatically_expires_abandoned_session_and_releases_gate(tmp_path):
    gate = NetworkMeasurementGate()
    manager = SustainedCheckManager(
        log_path=tmp_path / "data" / "network_check_session_log.csv",
        results_root=tmp_path / "data" / "network_check_results",
        settings=SustainedCheckSettings(session_ttl_seconds=0.05, download_chunk_bytes=1024),
        measurement_gate=gate,
    )
    session = manager.start_session(
        client_ip="10.0.0.10",
        direction="upload",
        duration_seconds=10,
        stream_count=1,
    )

    deadline = time.perf_counter() + 1
    while manager.active_session is not None and time.perf_counter() < deadline:
        time.sleep(0.01)

    assert manager.active_session is None
    assert gate.is_available() is True
    expired_path = tmp_path / "data" / "network_check_results" / f"{session.session_id}.json"
    assert json.loads(expired_path.read_text(encoding="utf-8"))["status"] == "failure"


@pytest.fixture()
def sustained_client(tmp_path):
    app = create_app(write_config(tmp_path))
    app.config.update(TESTING=True)
    return app.test_client(), tmp_path


def test_sustained_routes_validate_and_cancel(sustained_client):
    client, tmp_path = sustained_client
    invalid = client.post(
        "/network-check/sustained/sessions",
        json={"direction": "upload", "duration_seconds": 5, "stream_count": 1},
    )
    assert invalid.status_code == 400

    parallel = client.post(
        "/network-check/sustained/sessions",
        json={"direction": "upload", "duration_seconds": 10, "stream_count": 4},
    )
    assert parallel.status_code == 400
    assert "1개 연결" in parallel.json["error"]

    started = client.post(
        "/network-check/sustained/sessions",
        json={"direction": "upload", "duration_seconds": 10, "stream_count": 1},
    )
    assert started.status_code == 200
    session_id = started.json["session_id"]

    conflict = client.post(
        "/network-check/sustained/sessions",
        json={"direction": "download", "duration_seconds": 10, "stream_count": 1},
    )
    assert conflict.status_code == 409

    phase = client.post(
        f"/network-check/sustained/sessions/{session_id}/phase",
        json={"direction": "upload", "phase": "warmup"},
    )
    assert phase.status_code == 200
    uploaded = client.post(
        f"/network-check/sustained/sessions/{session_id}/upload/0",
        data=b"x" * 1024,
        content_type="application/octet-stream",
    )
    assert uploaded.status_code == 200

    cancelled = client.post(
        f"/network-check/sustained/sessions/{session_id}/cancel",
        json={"latency_samples_ms": [1.2, 1.4, 1.3]},
    )
    assert cancelled.status_code == 200
    assert cancelled.json["status"] == "cancelled"
    assert cancelled.json["excel_url"] == f"/network-check/sustained/results/{session_id}.xlsx"
    assert (tmp_path / "data" / "network_check_results" / f"{session_id}.json").exists()


def test_sustained_result_download_is_limited_to_origin_ip(sustained_client):
    client, _ = sustained_client
    started = client.post(
        "/network-check/sustained/sessions",
        json={"direction": "upload", "duration_seconds": 10, "stream_count": 1},
        environ_overrides={"REMOTE_ADDR": "10.0.0.10"},
    )
    session_id = started.json["session_id"]
    client.post(
        f"/network-check/sustained/sessions/{session_id}/cancel",
        json={},
        environ_overrides={"REMOTE_ADDR": "10.0.0.10"},
    )

    allowed = client.get(
        f"/network-check/sustained/results/{session_id}.json",
        environ_overrides={"REMOTE_ADDR": "10.0.0.10"},
    )
    denied = client.get(
        f"/network-check/sustained/results/{session_id}.json",
        environ_overrides={"REMOTE_ADDR": "10.0.0.11"},
    )
    excel_allowed = client.get(
        f"/network-check/sustained/results/{session_id}.xlsx",
        environ_overrides={"REMOTE_ADDR": "10.0.0.10"},
    )
    excel_denied = client.get(
        f"/network-check/sustained/results/{session_id}.xlsx",
        environ_overrides={"REMOTE_ADDR": "10.0.0.11"},
    )

    assert allowed.status_code == 200
    assert allowed.json["session_id"] == session_id
    assert denied.status_code == 403
    assert excel_allowed.status_code == 200
    assert excel_allowed.mimetype == EXCEL_MIME_TYPE
    assert "no-store" in excel_allowed.headers["Cache-Control"]
    assert "202607" in excel_allowed.headers["Content-Disposition"]
    assert session_id[:8] not in excel_allowed.headers["Content-Disposition"]
    assert load_workbook(BytesIO(excel_allowed.data)).sheetnames == ["결과 요약", "속도 변화"]
    assert excel_denied.status_code == 403


def test_sustained_excel_download_handles_missing_and_corrupt_results(sustained_client):
    client, tmp_path = sustained_client
    missing_id = "f" * 32
    missing = client.get(f"/network-check/sustained/results/{missing_id}.xlsx")

    corrupt_id = "e" * 32
    corrupt_path = tmp_path / "data" / "network_check_results" / f"{corrupt_id}.json"
    corrupt_path.write_text("{not-json", encoding="utf-8")
    corrupt = client.get(f"/network-check/sustained/results/{corrupt_id}.xlsx")

    assert missing.status_code == 404
    assert corrupt.status_code == 500


def test_sustained_csv_header_is_stable(sustained_client):
    _, tmp_path = sustained_client
    with (tmp_path / "data" / "network_check_session_log.csv").open(
        "r", encoding="utf-8-sig", newline=""
    ) as handle:
        assert next(csv.reader(handle)) == SUSTAINED_LOG_FIELDS
