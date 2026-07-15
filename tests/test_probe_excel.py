from __future__ import annotations

import json
from io import BytesIO

from flask import Flask
from openpyxl import load_workbook

from network_probe.excel import (
    EXCEL_MIME_TYPE,
    build_probe_excel,
    build_probe_excel_filename,
)
from network_probe.routes import create_probe_blueprint
from tests.test_network_probe import build_service


SESSION_ID = "0123456789abcdef0123456789abcdef"


def sample_probe_result() -> dict:
    def side(role: str, multiplier: int, *, telemetry: bool) -> dict:
        intervals = [
            {"index": 1, "bytes": 1_000_000 * multiplier, "mbps": 8.0 * multiplier},
            {"index": 2, "bytes": 2_000_000 * multiplier, "mbps": 16.0 * multiplier},
        ]
        stream_telemetry = {
            "available": True,
            "rtt_us": 1250,
            "min_rtt_us": 900,
            "cwnd_bytes": 262_144,
            "bytes_retrans": 4096,
            "fast_retransmits": 1,
            "duplicate_acks": 2,
            "timeout_episodes": 0,
        } if telemetry else {"available": False, "error": "수신 통계 없음"}
        return {
            "role": role,
            "bytes": 3_000_000 * multiplier,
            "duration_seconds": 2.0,
            "average_mbps": 12.0 * multiplier,
            "median_mbps": 12.0 * multiplier,
            "min_mbps": 8.0 * multiplier,
            "max_mbps": 16.0 * multiplier,
            "intervals": intervals,
            "streams": [{
                "stream_id": 0,
                "role": role,
                "bytes": 3_000_000 * multiplier,
                "duration_seconds": 2.0,
                "mbps": 12.0 * multiplier,
                "interval_bytes": [1_000_000 * multiplier, 2_000_000 * multiplier],
                "telemetry": stream_telemetry,
            }],
            "telemetry": stream_telemetry,
        }

    return {
        "schema_version": 1,
        "session_id": SESSION_ID,
        "started_at": "2026-07-14 14:30:00 +0900",
        "completed_at": "2026-07-14 14:31:06 +0900",
        "agent": {
            "agent_id": "abcdef0123456789abcdef0123456789",
            "hostname": "=CMD|' /C calc'!A0",
            "client_ip": "10.0.0.20",
        },
        "server_host": "10.0.0.10",
        "requested": {
            "direction": "full",
            "duration_seconds": 30,
            "warmup_seconds": 3.0,
            "stream_count": 1,
        },
        "phases": {
            "upload": {"sender": side("sender", 1, telemetry=True), "receiver": side("receiver", 1, telemetry=False)},
            "download": {"sender": side("sender", 2, telemetry=True), "receiver": side("receiver", 2, telemetry=False)},
        },
        "status": "completed",
        "error": "@SUM(A1:A2)",
    }


def test_probe_excel_has_summary_intervals_streams_and_charts():
    workbook = load_workbook(BytesIO(build_probe_excel(sample_probe_result())), data_only=False)

    assert workbook.sheetnames == ["결과 요약", "속도 변화", "기술 상세"]
    summary = workbook["결과 요약"]
    intervals = workbook["속도 변화"]
    streams = workbook["기술 상세"]
    assert summary["A1"].value == "TCP 전송 측정 결과"
    assert summary["A2"].value == "시간 기준: 한국 표준시(KST)"
    assert summary["B4"].value == "완료"
    assert summary["D4"].value == "전체"
    assert summary["F4"].value == 1
    assert summary["B5"].value.strftime("%Y-%m-%d %H:%M:%S") == "2026-07-14 14:30:00"
    assert summary["B6"].value.startswith("'=")
    assert summary["B7"].value.startswith("'@")
    assert summary["A11"].value == "업로드"
    assert summary["B11"].value == "측정 PC → 서버"
    assert summary["C11"].value == 12.0
    assert summary["D11"].value == 1.5
    assert summary["G11"].value == 1.25
    assert round(summary["H11"].value, 3) == 0.137
    assert summary["C12"].value == 24.0
    assert "정상·비정상을 자동 판정하지 않습니다" in summary["A15"].value
    assert intervals["A4"].value == "업로드"
    assert intervals["C4"].value == 8.0
    assert intervals["C3"].value == "속도(Mbps)"
    assert intervals["D3"].value == "초당 전송량(MB/s)"
    assert len(intervals._charts) == 2
    assert all(len(chart.series) == 3 for chart in intervals._charts)
    assert all(chart.y_axis.scaling.min == 0 for chart in intervals._charts)
    assert streams["A4"].value == "업로드"
    assert streams["B4"].value == "송신 측"
    assert streams["J4"].value == 1.25
    assert streams["I5"].value == "운영체제에서 제공하지 않음"
    assert all(
        cell.value != SESSION_ID
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
    )


def test_probe_excel_supports_single_direction_missing_telemetry_and_zero_bytes():
    result = sample_probe_result()
    result["requested"]["direction"] = "upload"
    result["phases"] = {"upload": result["phases"]["upload"]}
    sender = result["phases"]["upload"]["sender"]
    sender["telemetry"] = {"available": False, "error": "지원 안 됨"}

    workbook = load_workbook(BytesIO(build_probe_excel(result)))
    summary = workbook["결과 요약"]

    assert summary["A11"].value == "업로드"
    assert summary["A12"].value == "다운로드"
    assert summary["C12"].value == "측정 안 함"
    assert summary["G11"].value == "운영체제에서 제공하지 않음"
    assert summary["H11"].value == "운영체제에서 제공하지 않음"

    sender["telemetry"] = {"available": True, "bytes_retrans": 0, "rtt_us": 0, "min_rtt_us": 0}
    sender["bytes"] = 0
    zero_workbook = load_workbook(BytesIO(build_probe_excel(result)))
    assert zero_workbook["결과 요약"]["H11"].value == "계산 불가"


def test_probe_excel_supports_failed_result_without_completed_phases():
    result = sample_probe_result()
    result["status"] = "failed"
    result["error"] = "연결 실패"
    result["phases"] = {}

    workbook = load_workbook(BytesIO(build_probe_excel(result)))
    summary = workbook["결과 요약"]

    assert summary["B4"].value == "실패"
    assert summary["B7"].value == "연결 실패"
    assert summary["C11"].value == "측정 안 함"
    assert summary["C12"].value == "측정 안 함"


def test_probe_excel_filename_uses_kst_completion_time_without_session():
    assert build_probe_excel_filename(sample_probe_result()) == "TCP_전송측정_20260714_143106.xlsx"


def test_probe_excel_converts_offset_time_to_kst():
    result = sample_probe_result()
    result["started_at"] = "2026-07-14 05:30:00 +0000"
    result["completed_at"] = "2026-07-14 06:31:06 +0000"

    workbook = load_workbook(BytesIO(build_probe_excel(result)))

    assert workbook["결과 요약"]["B5"].value.strftime("%Y-%m-%d %H:%M:%S") == "2026-07-14 14:30:00"
    assert build_probe_excel_filename(result) == "TCP_전송측정_20260714_153106.xlsx"


def test_probe_excel_route_handles_saved_missing_and_corrupt_results(tmp_path):
    service, _ = build_service(tmp_path, enabled=False)
    app = Flask(__name__)
    app.register_blueprint(create_probe_blueprint(service))
    client = app.test_client()
    result_path = service.config.results_root / f"{SESSION_ID}.json"
    result_path.write_text(json.dumps(sample_probe_result(), ensure_ascii=False), encoding="utf-8")

    response = client.get(f"/api/network-probe/results/{SESSION_ID}.xlsx")

    assert response.status_code == 200
    assert response.mimetype == EXCEL_MIME_TYPE
    assert response.headers["Cache-Control"] == "no-store"
    assert response.headers["X-Content-Type-Options"] == "nosniff"
    assert "20260714_143106" in response.headers["Content-Disposition"]
    assert "01234567" not in response.headers["Content-Disposition"]
    assert load_workbook(BytesIO(response.data)).sheetnames == ["결과 요약", "속도 변화", "기술 상세"]

    missing = client.get(f"/api/network-probe/results/{'f' * 32}.xlsx")
    assert missing.status_code == 404

    corrupt_id = "e" * 32
    (service.config.results_root / f"{corrupt_id}.json").write_text("{not-json", encoding="utf-8")
    corrupt = client.get(f"/api/network-probe/results/{corrupt_id}.xlsx")
    assert corrupt.status_code == 500
