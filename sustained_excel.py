from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from excel_report import (
    EXCEL_DATETIME_FORMAT,
    KST_NOTE,
    add_speed_chart,
    kst_filename_timestamp,
    to_kst_excel_datetime,
)


EXCEL_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

_TITLE_FILL = PatternFill("solid", fgColor="1F4E3D")
_SECTION_FILL = PatternFill("solid", fgColor="DCE9E2")
_HEADER_FILL = PatternFill("solid", fgColor="E9EFEA")
_STATUS_FILLS = {
    "success": PatternFill("solid", fgColor="D9EAD3"),
    "failure": PatternFill("solid", fgColor="F4CCCC"),
    "cancelled": PatternFill("solid", fgColor="FFF2CC"),
}
_THIN_BORDER = Border(
    left=Side(style="thin", color="C8D0CA"),
    right=Side(style="thin", color="C8D0CA"),
    top=Side(style="thin", color="C8D0CA"),
    bottom=Side(style="thin", color="C8D0CA"),
)
_FORMULA_PREFIXES = ("=", "+", "-", "@")
_DIRECTION_LABELS = {"upload": "업로드", "download": "다운로드", "full": "전체"}
_STATUS_LABELS = {"success": "성공", "failure": "실패", "cancelled": "취소"}


class SustainedExcelError(RuntimeError):
    pass


def safe_excel_text(value: Any) -> str:
    text = ILLEGAL_CHARACTERS_RE.sub("", "" if value is None else str(value))[:32767]
    if text.lstrip().startswith(_FORMULA_PREFIXES):
        return f"'{text}"
    return text


def build_sustained_excel_filename(result: dict[str, Any]) -> str:
    timestamp = kst_filename_timestamp(result.get("completed_at"))
    return f"HTTP_시간측정_{timestamp}.xlsx"


def _as_mapping(value: Any) -> dict[str, Any]:
    return value if isinstance(value, dict) else {}


def _as_number(value: Any) -> float | None:
    if isinstance(value, bool):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if number != number or number in (float("inf"), float("-inf")):
        return None
    return number


def _as_integer(value: Any) -> int | None:
    number = _as_number(value)
    return int(number) if number is not None else None


def _style_title(sheet, title: str, *, end_column: int) -> None:
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    cell = sheet.cell(1, 1, safe_excel_text(title))
    cell.fill = _TITLE_FILL
    cell.font = Font(color="FFFFFF", bold=True, size=16)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 28


def _style_note(sheet, text: str, *, end_column: int) -> None:
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_column)
    cell = sheet.cell(2, 1, safe_excel_text(text))
    cell.font = Font(color="626B5F", italic=True, size=10)
    cell.alignment = Alignment(horizontal="right", vertical="center")


def _style_section(sheet, row: int, title: str, *, end_column: int) -> None:
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_column)
    cell = sheet.cell(row, 1, safe_excel_text(title))
    cell.fill = _SECTION_FILL
    cell.font = Font(bold=True, color="1F3D32")
    cell.alignment = Alignment(vertical="center")


def _style_header_row(sheet, row: int, start_column: int, end_column: int) -> None:
    for column in range(start_column, end_column + 1):
        cell = sheet.cell(row, column)
        cell.fill = _HEADER_FILL
        cell.font = Font(bold=True, color="27362F")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER


def _set_label_value(sheet, row: int, label_column: int, label: str, value: Any) -> None:
    label_cell = sheet.cell(row, label_column, safe_excel_text(label))
    value_cell = sheet.cell(row, label_column + 1, value)
    label_cell.fill = _HEADER_FILL
    label_cell.font = Font(bold=True, color="27362F")
    label_cell.border = _THIN_BORDER
    value_cell.border = _THIN_BORDER
    value_cell.alignment = Alignment(vertical="top", wrap_text=True)
    if isinstance(value, datetime):
        value_cell.number_format = EXCEL_DATETIME_FORMAT


def _time_value(value: Any) -> datetime | str:
    parsed = to_kst_excel_datetime(value)
    return parsed if parsed is not None else safe_excel_text(value)


def _build_summary_sheet(workbook: Workbook, result: dict[str, Any]) -> None:
    sheet = workbook.active
    sheet.title = "결과 요약"
    sheet.sheet_view.showGridLines = False
    _style_title(sheet, "HTTP 시간 측정 결과", end_column=8)
    _style_note(sheet, KST_NOTE, end_column=8)

    requested = _as_mapping(result.get("requested"))
    latency = _as_mapping(result.get("http_latency"))
    directions = _as_mapping(result.get("directions"))
    status = safe_excel_text(result.get("status", ""))

    _style_section(sheet, 3, "측정 정보", end_column=8)
    _set_label_value(sheet, 4, 1, "상태", safe_excel_text(_STATUS_LABELS.get(status, status)))
    sheet["B4"].fill = _STATUS_FILLS.get(status, PatternFill(fill_type=None))
    _set_label_value(
        sheet,
        4,
        3,
        "방향",
        safe_excel_text(_DIRECTION_LABELS.get(str(requested.get("direction", "")), requested.get("direction", ""))),
    )
    _set_label_value(sheet, 4, 5, "측정 PC IP", safe_excel_text(result.get("client_ip", "")))
    sheet.merge_cells("F4:H4")

    _set_label_value(sheet, 5, 1, "측정 시작", _time_value(result.get("started_at")))
    sheet.merge_cells("B5:C5")
    _set_label_value(sheet, 5, 4, "측정 완료", _time_value(result.get("completed_at")))
    sheet.merge_cells("E5:F5")

    _set_label_value(sheet, 6, 1, "측정 시간(초)", _as_number(requested.get("duration_seconds")))
    _set_label_value(sheet, 6, 3, "워밍업(초)", _as_number(requested.get("warmup_seconds")))
    _set_label_value(sheet, 6, 5, "HTTP 연결 수", _as_integer(requested.get("stream_count")))

    error = safe_excel_text(result.get("error", ""))
    if error:
        _set_label_value(sheet, 7, 1, "오류", error)
        sheet.merge_cells("B7:H7")
        sheet.row_dimensions[7].height = 30

    _style_section(sheet, 9, "핵심 결과", end_column=8)
    headers = (
        "방향",
        "평균 속도(Mbps)",
        "초당 전송량(MB/s)",
        "최저 속도(Mbps)",
        "최고 속도(Mbps)",
        "속도 변동(%)",
    )
    for column, header in enumerate(headers, start=1):
        sheet.cell(10, column, safe_excel_text(header))
    _style_header_row(sheet, 10, 1, len(headers))

    row = 11
    for direction in ("upload", "download"):
        if direction not in directions:
            continue
        summary = _as_mapping(directions.get(direction))
        average_mbps = _as_number(summary.get("average_mbps"))
        values = (
            _DIRECTION_LABELS[direction],
            average_mbps,
            average_mbps / 8 if average_mbps is not None else None,
            _as_number(summary.get("min_mbps")),
            _as_number(summary.get("max_mbps")),
            _as_number(summary.get("variability_percent")),
        )
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row, column, value)
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(vertical="center")
        for column in range(2, len(headers) + 1):
            sheet.cell(row, column).number_format = "#,##0.00"
        row += 1

    reference_row = max(14, row + 1)
    _style_section(sheet, reference_row, "참고", end_column=8)
    _set_label_value(sheet, reference_row + 1, 1, "HTTP 응답 시간(ms)", _as_number(latency.get("median_ms")))
    sheet.cell(reference_row + 1, 2).number_format = "#,##0.00"
    sheet.cell(
        reference_row + 1,
        3,
        "평균 속도는 전체 전송 성능, MB/s는 1초당 파일 용량입니다. 속도 변동은 낮을수록 일정합니다.",
    )
    sheet.merge_cells(start_row=reference_row + 1, start_column=3, end_row=reference_row + 1, end_column=8)
    sheet.cell(reference_row + 1, 3).alignment = Alignment(wrap_text=True, vertical="center")

    sheet.freeze_panes = "A11"
    if row > 11:
        sheet.auto_filter.ref = f"A10:F{row - 1}"
    widths = (16, 21, 21, 20, 20, 18, 14, 14)
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True


def _intervals(summary: Any) -> list[dict[str, Any]]:
    values = _as_mapping(summary).get("intervals", [])
    return [item for item in values if isinstance(item, dict)] if isinstance(values, list) else []


def _build_intervals_sheet(workbook: Workbook, result: dict[str, Any]) -> None:
    sheet = workbook.create_sheet("속도 변화")
    sheet.sheet_view.showGridLines = False
    _style_title(sheet, "1초 단위 HTTP 속도 변화", end_column=6)

    headers = ("방향", "시간(초)", "속도(Mbps)", "초당 전송량(MB/s)")
    for column, header in enumerate(headers, start=1):
        sheet.cell(3, column, safe_excel_text(header))
    _style_header_row(sheet, 3, 1, len(headers))

    directions = _as_mapping(result.get("directions"))
    row = 4
    chart_ranges: dict[str, tuple[int, int, list[int], list[float], int, int, float]] = {}
    for direction, average_column, label_column in (("upload", 5, 7), ("download", 6, 8)):
        direction_summary = _as_mapping(directions.get(direction))
        intervals = _intervals(direction_summary)
        start_row = row
        indexes: list[int] = []
        values: list[float] = []
        for fallback_index, item in enumerate(intervals, start=1):
            index = _as_integer(item.get("index")) or fallback_index
            mbps = _as_number(item.get("mbps"))
            if mbps is None or index <= 0:
                continue
            indexes.append(index)
            values.append(mbps)
            row_values = (_DIRECTION_LABELS[direction], index, mbps, mbps / 8)
            for column, value in enumerate(row_values, start=1):
                cell = sheet.cell(row, column, value)
                cell.border = _THIN_BORDER
                cell.alignment = Alignment(horizontal="right" if column > 1 else "center")
            sheet.cell(row, 3).number_format = "#,##0.00"
            sheet.cell(row, 4).number_format = "#,##0.00"
            row += 1
        if values:
            reported_average = _as_number(direction_summary.get("average_mbps"))
            average = reported_average if reported_average is not None else sum(values) / len(values)
            sheet.cell(3, average_column, f"평균 {average:.1f} Mbps")
            for average_row in range(start_row, row):
                sheet.cell(average_row, average_column, average)
            chart_ranges[direction] = (
                start_row,
                row - 1,
                indexes,
                values,
                average_column,
                label_column,
                average,
            )

    sheet.freeze_panes = "A4"
    sheet.auto_filter.ref = f"A3:D{max(3, row - 1)}"
    for column, width in enumerate((12, 13, 18, 22), start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.column_dimensions["E"].hidden = True
    sheet.column_dimensions["F"].hidden = True
    sheet.column_dimensions["G"].hidden = True
    sheet.column_dimensions["H"].hidden = True

    chart_number = 0
    for direction in ("upload", "download"):
        if direction not in chart_ranges:
            continue
        start_row, end_row, indexes, values, average_column, label_column, average = chart_ranges[direction]
        add_speed_chart(
            sheet,
            title=f"{_DIRECTION_LABELS[direction]} 속도 변화",
            header_row=3,
            start_row=start_row,
            end_row=end_row,
            time_column=2,
            speed_column=3,
            average_column=average_column,
            label_column=label_column,
            anchor="J3" if chart_number == 0 else "J21",
            color="246B54" if direction == "upload" else "C15F2E",
            indexes=indexes,
            values=values,
            average_value=average,
        )
        chart_number += 1

    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True


def build_sustained_excel(result: dict[str, Any]) -> bytes:
    if not isinstance(result, dict):
        raise SustainedExcelError("Excel로 변환할 측정 결과 형식이 올바르지 않습니다.")
    if not isinstance(result.get("directions"), dict):
        raise SustainedExcelError("측정 결과에 방향별 통계가 없습니다.")

    try:
        workbook = Workbook()
        workbook.properties.creator = "InternalUpload"
        workbook.properties.title = "HTTP 시간 측정 결과"
        _build_summary_sheet(workbook, result)
        _build_intervals_sheet(workbook, result)
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()
    except SustainedExcelError:
        raise
    except Exception as exc:
        raise SustainedExcelError("Excel 측정 결과를 생성할 수 없습니다.") from exc
