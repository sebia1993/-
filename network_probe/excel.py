from __future__ import annotations

import re
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from sustained_excel import EXCEL_MIME_TYPE, safe_excel_text


_TITLE_FILL = PatternFill("solid", fgColor="1F4E3D")
_SECTION_FILL = PatternFill("solid", fgColor="DCE9E2")
_HEADER_FILL = PatternFill("solid", fgColor="E9EFEA")
_STATUS_FILLS = {
    "completed": PatternFill("solid", fgColor="D9EAD3"),
    "failed": PatternFill("solid", fgColor="F4CCCC"),
    "cancelled": PatternFill("solid", fgColor="FFF2CC"),
}
_THIN_BORDER = Border(
    left=Side(style="thin", color="C8D0CA"),
    right=Side(style="thin", color="C8D0CA"),
    top=Side(style="thin", color="C8D0CA"),
    bottom=Side(style="thin", color="C8D0CA"),
)
_DIRECTION_LABELS = {"upload": "업로드", "download": "다운로드", "full": "전체"}
_STATUS_LABELS = {"completed": "완료", "failed": "실패", "cancelled": "취소"}
_NOT_MEASURED = "측정 안 함"
_TELEMETRY_UNAVAILABLE = "운영체제에서 제공하지 않음"


class ProbeExcelError(RuntimeError):
    pass


def _as_mapping(value: Any) -> dict[str, Any]:
    return value if isinstance(value, dict) else {}


def _as_number(value: Any) -> float | None:
    if isinstance(value, bool):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if number != number or number in (float("inf"), float("-inf")):
        return None
    return number


def _as_integer(value: Any) -> int | None:
    number = _as_number(value)
    return int(number) if number is not None else None


def _style_title(sheet, title: str, end_column: int) -> None:
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    cell = sheet.cell(1, 1, safe_excel_text(title))
    cell.fill = _TITLE_FILL
    cell.font = Font(color="FFFFFF", bold=True, size=16)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 28


def _style_section(sheet, row: int, title: str, end_column: int) -> None:
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_column)
    cell = sheet.cell(row, 1, safe_excel_text(title))
    cell.fill = _SECTION_FILL
    cell.font = Font(bold=True, color="1F3D32")


def _style_header(sheet, row: int, end_column: int) -> None:
    for column in range(1, end_column + 1):
        cell = sheet.cell(row, column)
        cell.fill = _HEADER_FILL
        cell.font = Font(bold=True, color="27362F")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER


def _set_label_value(sheet, row: int, column: int, label: str, value: Any) -> None:
    label_cell = sheet.cell(row, column, safe_excel_text(label))
    value_cell = sheet.cell(row, column + 1, value)
    label_cell.fill = _HEADER_FILL
    label_cell.font = Font(bold=True, color="27362F")
    label_cell.border = _THIN_BORDER
    value_cell.border = _THIN_BORDER
    value_cell.alignment = Alignment(vertical="top", wrap_text=True)


def _phase_endpoints(direction: str) -> tuple[str, str]:
    if direction == "upload":
        return "클라이언트", "서버"
    return "서버", "클라이언트"


def _phase_path(direction: str) -> str:
    return "측정 PC → 서버" if direction == "upload" else "서버 → 측정 PC"


def _telemetry_value(telemetry: dict[str, Any], key: str, *, milliseconds: bool = False) -> Any:
    if not telemetry.get("available"):
        return _TELEMETRY_UNAVAILABLE
    value = _as_number(telemetry.get(key))
    if value is None:
        return _TELEMETRY_UNAVAILABLE
    return value / 1000 if milliseconds else int(value)


def _format_bytes_for_summary(value: Any) -> str:
    byte_count = _as_number(value)
    if byte_count is None or byte_count < 0:
        return _TELEMETRY_UNAVAILABLE
    if byte_count >= 1024 * 1024:
        return f"{byte_count / 1024 / 1024:.1f} MB"
    if byte_count >= 1024:
        return f"{byte_count / 1024:.1f} KB"
    return f"{byte_count:.0f} B"


def _retransmission_rate(sender: dict[str, Any], telemetry: dict[str, Any]) -> float | str:
    if not telemetry.get("available"):
        return _TELEMETRY_UNAVAILABLE
    retransmitted = _as_number(telemetry.get("bytes_retrans"))
    sent = _as_number(sender.get("bytes"))
    if retransmitted is None:
        return _TELEMETRY_UNAVAILABLE
    if sent is None or sent <= 0:
        return "계산 불가"
    return retransmitted / sent * 100


def _direction_difference(phases: dict[str, Any]) -> float | None:
    upload = _as_number(_as_mapping(_as_mapping(phases.get("upload")).get("receiver")).get("average_mbps"))
    download = _as_number(_as_mapping(_as_mapping(phases.get("download")).get("receiver")).get("average_mbps"))
    if upload is None or download is None:
        return None
    maximum = max(upload, download)
    return abs(upload - download) / maximum * 100 if maximum > 0 else None


def build_probe_excel_filename(result: dict[str, Any]) -> str:
    completed_at = safe_excel_text(result.get("completed_at", ""))
    digits = "".join(re.findall(r"\d", completed_at))
    timestamp = f"{digits[:8]}-{digits[8:14]}" if len(digits) >= 14 else "unknown-time"
    session_id = safe_excel_text(result.get("session_id", ""))
    short_id = session_id[:8] if re.fullmatch(r"[0-9a-f]{8,32}", session_id) else "result"
    return f"tcp-probe_{timestamp}_{short_id}.xlsx"


def _build_summary_sheet(workbook: Workbook, result: dict[str, Any]) -> None:
    sheet = workbook.active
    sheet.title = "측정 요약"
    sheet.sheet_view.showGridLines = False
    _style_title(sheet, "TCP 전송 성능 측정 결과", 11)

    requested = _as_mapping(result.get("requested"))
    agent = _as_mapping(result.get("agent"))
    phases = _as_mapping(result.get("phases"))
    status = safe_excel_text(result.get("status", ""))

    _style_section(sheet, 3, "측정 정보", 11)
    _set_label_value(sheet, 4, 1, "측정 상태", safe_excel_text(_STATUS_LABELS.get(status, status)))
    sheet["B4"].fill = _STATUS_FILLS.get(status, PatternFill(fill_type=None))
    _set_label_value(sheet, 4, 3, "세션 ID", safe_excel_text(result.get("session_id", "")))
    sheet.merge_cells("D4:F4")
    _set_label_value(sheet, 4, 7, "측정 방향", safe_excel_text(_DIRECTION_LABELS.get(str(requested.get("direction", "")), requested.get("direction", ""))))
    _set_label_value(sheet, 4, 9, "TCP 스트림", _as_integer(requested.get("stream_count")))

    _set_label_value(sheet, 5, 1, "오류 내용", safe_excel_text(result.get("error", "")))
    sheet.merge_cells("B5:K5")
    sheet.row_dimensions[5].height = 34

    _set_label_value(sheet, 6, 1, "시작 시각", safe_excel_text(result.get("started_at", "")))
    sheet.merge_cells("B6:C6")
    _set_label_value(sheet, 6, 4, "완료 시각", safe_excel_text(result.get("completed_at", "")))
    sheet.merge_cells("E6:F6")
    _set_label_value(sheet, 6, 7, "본 측정(초)", _as_number(requested.get("duration_seconds")))
    _set_label_value(sheet, 6, 9, "워밍업(초)", _as_number(requested.get("warmup_seconds")))
    _set_label_value(sheet, 7, 1, "측정 PC", safe_excel_text(agent.get("hostname", "")))
    sheet.merge_cells("B7:C7")
    _set_label_value(sheet, 7, 4, "측정 PC IP", safe_excel_text(agent.get("client_ip", "")))
    sheet.merge_cells("E7:F7")
    _set_label_value(sheet, 7, 7, "서버 주소", safe_excel_text(result.get("server_host", "")))
    sheet.merge_cells("H7:K7")

    _style_section(sheet, 9, "핵심 결과", 11)
    headers = (
        "방향", "측정 경로", "실제 수신 평균(Mbps)", "파일 전송량(MB/s)",
        "중앙값(Mbps)", "최소(Mbps)", "최대(Mbps)", "RTT(ms)", "최소 RTT(ms)",
        "TCP 재전송량", "재전송률(%)",
    )
    for column, header in enumerate(headers, start=1):
        sheet.cell(10, column, safe_excel_text(header))
    _style_header(sheet, 10, 11)

    row = 11
    for direction in ("upload", "download"):
        if direction not in phases:
            values = (_DIRECTION_LABELS[direction], _phase_path(direction)) + (_NOT_MEASURED,) * 9
            for column, value in enumerate(values, start=1):
                cell = sheet.cell(row, column, value)
                cell.border = _THIN_BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            row += 1
            continue
        phase = _as_mapping(phases.get(direction))
        sender = _as_mapping(phase.get("sender"))
        receiver = _as_mapping(phase.get("receiver"))
        telemetry = _as_mapping(sender.get("telemetry"))
        average_mbps = _as_number(receiver.get("average_mbps"))
        values = (
            _DIRECTION_LABELS[direction], _phase_path(direction), average_mbps,
            average_mbps / 8 if average_mbps is not None else None,
            _as_number(receiver.get("median_mbps")), _as_number(receiver.get("min_mbps")),
            _as_number(receiver.get("max_mbps")),
            _telemetry_value(telemetry, "rtt_us", milliseconds=True),
            _telemetry_value(telemetry, "min_rtt_us", milliseconds=True),
            _format_bytes_for_summary(telemetry.get("bytes_retrans")) if telemetry.get("available") else _TELEMETRY_UNAVAILABLE,
            _retransmission_rate(sender, telemetry),
        )
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row, column, value)
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        for column in range(3, 8):
            sheet.cell(row, column).number_format = "#,##0.0"
        for column in (8, 9):
            sheet.cell(row, column).number_format = "#,##0.00"
        sheet.cell(row, 11).number_format = "0.000"
        row += 1

    comparison_row = row + 1
    _style_section(sheet, comparison_row, "업로드·다운로드 비교", 11)
    difference = _direction_difference(phases)
    _set_label_value(
        sheet,
        comparison_row + 1,
        1,
        "실제 수신 평균 속도 차이",
        difference if difference is not None else _NOT_MEASURED,
    )
    sheet.cell(comparison_row + 1, 2).number_format = "0.000\"%\""
    sheet.merge_cells(start_row=comparison_row + 1, start_column=2, end_row=comparison_row + 1, end_column=4)
    sheet.cell(comparison_row + 1, 5, "양방향 속도 중 큰 값을 기준으로 계산하며 정상·비정상 판정값이 아닙니다.")
    sheet.merge_cells(start_row=comparison_row + 1, start_column=5, end_row=comparison_row + 1, end_column=11)
    sheet.cell(comparison_row + 1, 5).alignment = Alignment(wrap_text=True, vertical="top")

    guidance_row = comparison_row + 3
    _style_section(sheet, guidance_row, "결과 해석 참고", 11)
    guidance = (
        "실제 수신 평균 속도를 우선 확인하세요. 기준 속도가 설정되지 않아 정상·비정상을 자동 판정하지 않습니다. "
        "측정값에는 측정 PC와 서버의 CPU·NIC 성능, 무선 환경과 네트워크 경로 상태가 함께 반영됩니다."
    )
    sheet.cell(guidance_row + 1, 1, guidance)
    sheet.merge_cells(start_row=guidance_row + 1, start_column=1, end_row=guidance_row + 2, end_column=11)
    sheet.cell(guidance_row + 1, 1).alignment = Alignment(wrap_text=True, vertical="top")
    sheet.row_dimensions[guidance_row + 1].height = 34

    widths = (11, 20, 20, 19, 17, 15, 15, 13, 15, 18, 16)
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.freeze_panes = "A11"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True


def _interval_map(side: Any) -> dict[int, dict[str, Any]]:
    intervals = _as_mapping(side).get("intervals", [])
    if not isinstance(intervals, list):
        return {}
    mapped: dict[int, dict[str, Any]] = {}
    for fallback_index, item in enumerate(intervals, start=1):
        if not isinstance(item, dict):
            continue
        index = _as_integer(item.get("index")) or fallback_index
        if index > 0:
            mapped[index] = item
    return mapped


def _build_interval_sheet(workbook: Workbook, result: dict[str, Any]) -> None:
    sheet = workbook.create_sheet("구간별 속도")
    sheet.sheet_view.showGridLines = False
    _style_title(sheet, "1초 구간별 TCP 처리량", 6)
    headers = (
        "방향", "구간(초)", "송신 측 기록(Byte)", "송신 측 기록(Mbps)",
        "수신 측 실제(Byte)", "수신 측 실제(Mbps)",
    )
    for column, header in enumerate(headers, start=1):
        sheet.cell(3, column, safe_excel_text(header))
    _style_header(sheet, 3, 6)

    phases = _as_mapping(result.get("phases"))
    row = 4
    chart_ranges: dict[str, tuple[int, int]] = {}
    for direction in ("upload", "download"):
        phase = _as_mapping(phases.get(direction))
        sender_intervals = _interval_map(phase.get("sender"))
        receiver_intervals = _interval_map(phase.get("receiver"))
        indexes = sorted(set(sender_intervals) | set(receiver_intervals))
        start_row = row
        for index in indexes:
            sender = sender_intervals.get(index, {})
            receiver = receiver_intervals.get(index, {})
            values = (
                _DIRECTION_LABELS[direction], index,
                _as_integer(sender.get("bytes")), _as_number(sender.get("mbps")),
                _as_integer(receiver.get("bytes")), _as_number(receiver.get("mbps")),
            )
            for column, value in enumerate(values, start=1):
                cell = sheet.cell(row, column, value)
                cell.border = _THIN_BORDER
                cell.alignment = Alignment(horizontal="right" if column > 1 else "center")
            for column in (3, 5):
                sheet.cell(row, column).number_format = "#,##0"
            for column in (4, 6):
                sheet.cell(row, column).number_format = "#,##0.00"
            row += 1
        if row > start_row:
            chart_ranges[direction] = (start_row, row - 1)

    sheet.freeze_panes = "A4"
    sheet.auto_filter.ref = f"A3:F{max(3, row - 1)}"
    for column, width in enumerate((12, 13, 18, 17, 18, 17), start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width

    for chart_index, direction in enumerate(("upload", "download")):
        if direction not in chart_ranges:
            continue
        start_row, end_row = chart_ranges[direction]
        chart = LineChart()
        chart.title = f"{_DIRECTION_LABELS[direction]} 송신 측 기록·수신 측 실제 속도"
        chart.y_axis.title = "Mbps"
        chart.x_axis.title = "구간(초)"
        chart.height = 8
        chart.width = 16
        chart.legend.position = "b"
        categories = Reference(sheet, min_col=2, min_row=start_row, max_row=end_row)
        for column in (4, 6):
            data = Reference(sheet, min_col=column, min_row=3, max_row=end_row)
            chart.add_data(data, titles_from_data=True, from_rows=False)
            if start_row > 4:
                chart.series[-1].val.numRef.f = f"'{sheet.title}'!${get_column_letter(column)}${start_row}:${get_column_letter(column)}${end_row}"
        chart.set_categories(categories)
        sheet.add_chart(chart, "H3" if chart_index == 0 else "H20")

    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True


def _build_stream_sheet(workbook: Workbook, result: dict[str, Any]) -> None:
    sheet = workbook.create_sheet("스트림 상세")
    sheet.sheet_view.showGridLines = False
    headers = (
        "방향", "역할", "측정 지점", "스트림 ID", "전송량(Byte)", "전송량(MiB)", "시간(초)",
        "평균 속도(Mbps)", "TCP 상세 통계", "왕복 지연 RTT(ms)", "최소 왕복 지연(ms)",
        "송신 혼잡 윈도우(Byte)", "TCP 재전송(Byte)", "빠른 재전송(회)",
        "중복 ACK(회)", "타임아웃(회)", "비고",
    )
    _style_title(sheet, "TCP 전송 성능 스트림 상세", len(headers))
    for column, header in enumerate(headers, start=1):
        sheet.cell(3, column, safe_excel_text(header))
    _style_header(sheet, 3, len(headers))

    phases = _as_mapping(result.get("phases"))
    row = 4
    for direction in ("upload", "download"):
        phase = _as_mapping(phases.get(direction))
        sender_endpoint, receiver_endpoint = _phase_endpoints(direction)
        for role, endpoint in (("sender", sender_endpoint), ("receiver", receiver_endpoint)):
            streams = _as_mapping(phase.get(role)).get("streams", [])
            if not isinstance(streams, list):
                continue
            for stream in streams:
                if not isinstance(stream, dict):
                    continue
                telemetry = _as_mapping(stream.get("telemetry"))
                available = bool(telemetry.get("available"))
                byte_count = _as_integer(stream.get("bytes"))
                values = (
                    _DIRECTION_LABELS[direction], "송신 측" if role == "sender" else "수신 측", endpoint,
                    _as_integer(stream.get("stream_id")), byte_count,
                    byte_count / (1024 * 1024) if byte_count is not None else None,
                    _as_number(stream.get("duration_seconds")), _as_number(stream.get("mbps")),
                    "제공" if available else _TELEMETRY_UNAVAILABLE,
                    _telemetry_value(telemetry, "rtt_us", milliseconds=True),
                    _telemetry_value(telemetry, "min_rtt_us", milliseconds=True),
                    _telemetry_value(telemetry, "cwnd_bytes"), _telemetry_value(telemetry, "bytes_retrans"),
                    _telemetry_value(telemetry, "fast_retransmits"), _telemetry_value(telemetry, "duplicate_acks"),
                    _telemetry_value(telemetry, "timeout_episodes"),
                    "" if available else safe_excel_text(telemetry.get("error", "TCP 상세 통계를 사용할 수 없습니다.")),
                )
                for column, value in enumerate(values, start=1):
                    cell = sheet.cell(row, column, value)
                    cell.border = _THIN_BORDER
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                for column in (4, 5, 12, 13, 14, 15, 16):
                    sheet.cell(row, column).number_format = "#,##0"
                for column in (6, 7, 8, 10, 11):
                    sheet.cell(row, column).number_format = "#,##0.00"
                row += 1

    sheet.freeze_panes = "A4"
    sheet.auto_filter.ref = f"A3:Q{max(3, row - 1)}"
    widths = (11, 10, 14, 12, 18, 16, 13, 16, 13, 13, 15, 20, 18, 15, 13, 13, 34)
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True


def build_probe_excel(result: dict[str, Any]) -> bytes:
    if not isinstance(result, dict):
        raise ProbeExcelError("Excel로 변환할 TCP 측정 결과 형식이 올바르지 않습니다.")
    if not isinstance(result.get("phases"), dict):
        raise ProbeExcelError("TCP 측정 결과에 방향별 통계가 없습니다.")
    try:
        workbook = Workbook()
        workbook.properties.creator = "InternalUpload"
        workbook.properties.title = "TCP 전송 성능 측정 결과"
        _build_summary_sheet(workbook, result)
        _build_interval_sheet(workbook, result)
        _build_stream_sheet(workbook, result)
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()
    except ProbeExcelError:
        raise
    except Exception as exc:
        raise ProbeExcelError("TCP Excel 측정 결과를 생성할 수 없습니다.") from exc


__all__ = [
    "EXCEL_MIME_TYPE",
    "ProbeExcelError",
    "build_probe_excel",
    "build_probe_excel_filename",
]
